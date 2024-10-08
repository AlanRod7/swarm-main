import asyncio
import hashlib
import re
from array import array
from datetime import datetime, timedelta
from functools import wraps

import jwt
import mysql.connector
import numpy as np
from flask import (Flask, jsonify, make_response, render_template, request,
                   send_from_directory, session)
from flask_sqlalchemy import SQLAlchemy
from numpy import ndarray
from openpyxl import load_workbook

from aco import ejecutar_aco
from ba import ejecutar_ba
from da import ejecutar_da
from daaco import ejecutar_daaco
from daba import ejecutar_daba
from dapso import ejecutar_dapso
from mooraaco import ejecutar_mooraaco
from mooraba import ejecutar_mooraba
from moorapso import ejecutar_moorapso
from moorav import ejecutar_moorav
#importacion de algoritmos
from pso import ejecutar_pso
from topsis import ejecutar_topsis
from topsisaco import ejecutar_topsisaco
from topsisba import ejecutar_topsisba
from topsispso import ejecutar_topsispso

db = SQLAlchemy()
app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = "postgresql://postgres:53Zc56erWZsY7@db.ogyulszumjcstdztensn.supabase.co:5432/postgres"
app.config['SECRET_KEY'] = '563cebb3aceb49e0a6c79ded5c717235'

db.init_app(app)
class SupaUser(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombredb = db.Column(db.String, unique=True, nullable=False)
    escolaridaddb = db.Column(db.String, nullable=False)
    #nacimientodb = db.Column(db.Date, nullable=False)
    emaildb = db.Column(db.String, nullable=False)
    passworddb = db.Column(db.String, nullable=False)
    
                    #Acerca de
#-------------------------------------------------------------------------------------------------------------------
@app.route('/acercade')
def acercade():
    return render_template('acercade.html')
@app.route('/casoexperimental')
def casoexperimental():
    return render_template('casoexperimental.html')
                #Algoritmos PSO
#-------------------------------------------------------------------------------------------------------------------
@app.route('/pso')
def pso():
    try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]
        
        # Llama a la función de procesar_datos en pso.py
        datosPso = asyncio.run(ejecutar_pso(w, wwi, c1, c2, T, r1, r2))

        return render_template('pso.html', datosPso=datosPso)
    except Exception as e:
        return render_template('pso.html', error_message=str(e))


@app.route('/pso', methods=['POST'])
def calcular_pso():
    try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Llama a la función de PSO en pso.py
        datosPso = asyncio.run(ejecutar_pso(w, wwi, c1, c2, T, r1, r2))
        print("Resultados de la ejecución:", datosPso)

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosPso)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_pso: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------
@app.route('/dapso')
def dapso():
    try:
        # Obtener los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]
        
        # Llamar a la función de procesar_datos en pso.py
        datosDapso = asyncio.run(ejecutar_dapso(w, wwi, c1, c2, T, r1, r2))

        return render_template('dapso.html', datosDapso=datosDapso)
    except Exception as e:
        return render_template('dapso.html', error_message=str(e))

@app.route('/dapso', methods=['POST'])
def calcular_dapso():
   try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Llama a la función de PSO en pso.py
        datosDapso = asyncio.run(ejecutar_dapso(w, wwi, c1, c2, T, r1, r2))
        print("Resultados de la ejecución:", datosDapso)

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosDapso)
   except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
        print(f'Error en calcular_dapso: {str(e)}')
   return jsonify({'error': 'Ocurrió un error en el servidor'}), 500

#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------

@app.route('/moorapso', methods=['POST'])
def calcular_moorapso():
    try:
        # Obtén los datos del formulario
        w_input = [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Llama a la función de PSO en pso.py
        datosMoorapso = asyncio.run(ejecutar_moorapso(w, wwi, c1, c2, T, r1, r2))
        print("Resultados de la ejecución:", datosMoorapso)

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosMoorapso)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_moorapso: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500


@app.route('/moorapso')
def moorapso():
    try:
        # Obtén los datos del formulario
        w_input = [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]
        
        # Llama a la función de procesar_datos en pso.py
        datosMoorapso = asyncio.run(ejecutar_moorapso(w, wwi, c1, c2, T, r1, r2))

        return render_template('moorapso.html', datosMoorapso=datosMoorapso)
    except Exception as e:
        return render_template('moorapso.html', error_message=str(e))
    
    #-------------------------------------------------------------------------------------------------------------------

    #-------------------------------------------------------------------------------------------------------------------

@app.route('/topsispso')
def topsispso():
     try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]
        
        # Llama a la función de procesar_datos en pso.py
        datosTopsispso = asyncio.run(ejecutar_topsispso(w, wwi, c1, c2, T, r1, r2))

        return render_template('topsispso.html', datosTopsispso=datosTopsispso)
     except Exception as e:
        return render_template('topsispso.html', error_message=str(e))

@app.route('/topsispso', methods=['POST'])
def calcular_topsispso():
    try:
    # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Llama a la función de PSO en pso.py
        datosTopsispso = asyncio.run(ejecutar_topsispso(w, wwi, c1, c2, T, r1, r2))
        print("Resultados de la ejecución:", datosTopsispso)

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosTopsispso)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_topsispso: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500

#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------


@app.route('/comparacion')
def comparacionPura():
    try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]
        
        # Llama a la función de procesar_datos en pso.py
        datosPso = asyncio.run(ejecutar_pso(w, wwi, c1, c2, T, r1, r2))
        datosDapso = asyncio.run(ejecutar_dapso(w, wwi, c1, c2, T, r1, r2))
        datosMoorapso = asyncio.run(ejecutar_moorapso(w, wwi, c1, c2, T, r1, r2))
        datosTopsispso = asyncio.run(ejecutar_topsispso(w, wwi, c1, c2, T, r1, r2))

        return render_template('comparacion.html', datosPso=datosPso, datosDapso = datosDapso , datosMoorapso = datosMoorapso, datosTopsispso = datosTopsispso)
    except Exception as e:
        return render_template('comparacion.html', error_message=str(e))


@app.route('/comparacion', methods=['POST'])
def calcular_comparacionPura():
    try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Llama a la función de PSO en pso.py
        datosPso = asyncio.run(ejecutar_pso(w, wwi, c1, c2, T, r1, r2))
        datosDapso = asyncio.run(ejecutar_dapso(w, wwi, c1, c2, T, r1, r2))
        datosMoorapso = asyncio.run(ejecutar_moorapso(w, wwi, c1, c2, T, r1, r2))
        datosTopsispso = asyncio.run(ejecutar_topsispso(w, wwi, c1, c2, T, r1, r2))
        print("Resultados de la ejecución:", datosPso) 
        print("Resultados de la ejecución:", datosDapso) 
        print("Resultados de la ejecución:", datosMoorapso) 
        print("Resultados de la ejecución:", datosTopsispso) 

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosPso, datosDapso, datosMoorapso, datosTopsispso)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_comparacion: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500

#-------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------


@app.route('/comparacionPso')
def comparacion():
    try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]
        
        # Llama a la función de procesar_datos en pso.py
        datosPso = asyncio.run(ejecutar_pso(w, wwi, c1, c2, T, r1, r2))
        datosDapso = asyncio.run(ejecutar_dapso(w, wwi, c1, c2, T, r1, r2))
        datosMoorapso = asyncio.run(ejecutar_moorapso(w, wwi, c1, c2, T, r1, r2))
        datosTopsispso = asyncio.run(ejecutar_topsispso(w, wwi, c1, c2, T, r1, r2))

        return render_template('comparacionPso.html', datosPso=datosPso, datosDapso = datosDapso , datosMoorapso = datosMoorapso, datosTopsispso = datosTopsispso)
    except Exception as e:
        return render_template('comparacionPso.html', error_message=str(e))


@app.route('/comparacionPso', methods=['POST'])
def calcular_comparacion():
    try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Llama a la función de PSO en pso.py
        datosPso = asyncio.run(ejecutar_pso(w, wwi, c1, c2, T, r1, r2))
        datosDapso = asyncio.run(ejecutar_dapso(w, wwi, c1, c2, T, r1, r2))
        datosMoorapso = asyncio.run(ejecutar_moorapso(w, wwi, c1, c2, T, r1, r2))
        datosTopsispso = asyncio.run(ejecutar_topsispso(w, wwi, c1, c2, T, r1, r2))
        print("Resultados de la ejecución:", datosPso) 
        print("Resultados de la ejecución:", datosDapso) 
        print("Resultados de la ejecución:", datosMoorapso) 
        print("Resultados de la ejecución:", datosTopsispso) 

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosPso, datosDapso, datosMoorapso, datosTopsispso)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_comparacion: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500

#-------------------------------------------------------------------------------------------------------------------
            # Algoritmos BA
#-------------------------------------------------------------------------------------------------------------------
@app.route('/ba')
def ba():
    try:
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos en pso.py
        datosBa = asyncio.run(ejecutar_ba(w, alpha, gamma, iter_max))

        return render_template('ba.html', datosBa=datosBa)
    except Exception as e:
        return render_template('ba.html', error_message=str(e))


@app.route('/ba', methods=['POST'])
def calcular_ba():
    try:
        # Obtén los datos del formulario
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])

        # Llama a la función de PSO en pso.py
        datosBa = asyncio.run(ejecutar_ba(w, alpha, gamma, iter_max))
        print("Resultados de la ejecución:", datosBa)

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosBa)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_ba: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------
            
#-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta DA -BA

@app.route('/daba')
def daba():
    try:
        # Obtén los datos del formulario
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos en pso.py
        datosDaba = asyncio.run(ejecutar_daba(w, alpha, gamma, iter_max))

        return render_template('daba.html', datosDaba=datosDaba)
    except Exception as e:
        return render_template('daba.html', error_message=str(e))


@app.route('/daba', methods=['POST'])
def calcular_daba():
    try:
        # Obtén los datos del formulario de la solicitud POST
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])

        # Llama a la función de PSO en pso.py
        datosDaba = asyncio.run(ejecutar_daba(w, alpha, gamma, iter_max))
        print("Resultados de la ejecución:", datosDaba)

        # Devuelve los resultados como JSON
        return jsonify(datosDaba)
    except Exception as e:
        # Manejo de errores
        print(f'Error en calcular_mooraba: {str(e)}')
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta MOORA - BA

@app.route('/mooraba')
def mooraba():
    try:
        # Obtén los datos del formulario
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos en pso.py
        datosMooraba = asyncio.run(ejecutar_mooraba(w, alpha, gamma, iter_max))

        return render_template('mooraba.html', datosMooraba=datosMooraba)
    except Exception as e:
        return render_template('mooraba.html', error_message=str(e))


@app.route('/mooraba', methods=['POST'])
def calcular_mooraba():
    try:
        # Obtén los datos del formulario de la solicitud POST
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])

        # Llama a la función de PSO en pso.py
        datosMooraba = asyncio.run(ejecutar_mooraba(w, alpha, gamma, iter_max))
        print("Resultados de la ejecución:", datosMooraba)

        # Devuelve los resultados como JSON
        return jsonify(datosMooraba)
    except Exception as e:
        # Manejo de errores
        print(f'Error en calcular_mooraba: {str(e)}')
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------
@app.route('/topsisba')
def topsisba():
    try:
        # Obtén los datos del formulario
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos en pso.py
        datosTopsisba = asyncio.run(ejecutar_topsisba(w, alpha, gamma, iter_max))

        return render_template('topsisba.html', datosTopsisba=datosTopsisba)
    except Exception as e:
        return render_template('topsisba.html', error_message=str(e))


@app.route('/topsisba', methods=['POST'])
def calcular_topsisba():
    try:
        # Obtén los datos del formulario de la solicitud POST
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])

        # Llama a la función de PSO en pso.py
        datosTopsisba = asyncio.run(ejecutar_topsisba(w, alpha, gamma, iter_max))
        print("Resultados de la ejecución:", datosTopsisba)

        # Devuelve los resultados como JSON
        return jsonify(datosTopsisba)
    except Exception as e:
        # Manejo de errores
        print(f'Error en calcular_topsisBa: {str(e)}')
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------

#Algoritmos ACO

#-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta DA-ACO

@app.route('/daaco')
def daaco():
    try:
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])

        
        # Llama a la función de procesar_datos 
        datosDaaco = asyncio.run(ejecutar_daaco(w, alphaAco, beta, rho, Q, n_ants, iter_max))

        return render_template('daaco.html', datosDaaco=datosDaaco)
    except Exception as e:
        return render_template('daaco.html', error_message=str(e))


@app.route('/daaco', methods=['POST'])
def calcular_daaco():
    try:
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])

        # Llama a la función de PSO en pso.py
        datosDaaco = asyncio.run(ejecutar_daaco(w, alphaAco, beta, rho, Q, n_ants, iter_max))
        print("Resultados de la ejecución:", datosDaaco)

        # Devuelve los resultados como JSON
        return jsonify(datosDaaco)
    except Exception as e:
        # Manejo de errores
        print(f'Error en calcular_daaco: {str(e)}')
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta MOORA-ACO

@app.route('/mooraaco')
def mooraaco():
    try:
        ev_input = request.form['ev']  # Obtén el valor
        
        ev_values = ev_input.split(',')
        
        EV = [str(value) for value in ev_values if value.strip() != '']
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos 
        datosMooraaco = asyncio.run(ejecutar_mooraaco(EV,w,alphaAco,beta,rho,Q,n_ants,iter_max))

        return render_template('mooraaco.html', datosMooraaco=datosMooraaco)
    except Exception as e:
        return render_template('mooraaco.html', error_message=str(e))


@app.route('/mooraaco', methods=['POST'])
def calcular_mooraaco():
    try:
        ev_input = request.form['ev']  # Obtén el valor
        
        ev_values = ev_input.split(',')
        
        EV = [str(value) for value in ev_values if value.strip() != '']
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])

        # Llama a la función de PSO en pso.py
        datosMooraaco = asyncio.run(ejecutar_mooraaco(EV,w,alphaAco,beta,rho,Q,n_ants,iter_max))
        print("Resultados de la ejecución:", datosMooraaco)

        # Devuelve los resultados como JSON
        return jsonify(datosMooraaco)
    except Exception as e:
        # Manejo de errores
        print(f'Error en calcular_Mooraaco: {str(e)}')
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------
    
    #-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta TOPSIS-ACO

@app.route('/topsisaco')
def topsisaco():
    try:
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos

        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])
        
        # benefit_input = request.form['be']  # Obtén el valor
        
        # benefit_values = benefit_input.split(',')
        
        # benefit_attributes = [int(value) for value in benefit_values if value.strip() != '']
        
        # Llama a la función de procesar_datos 
        datosTopsisaco = asyncio.run(ejecutar_topsisaco(w,alphaAco,beta,rho,Q,n_ants,iter_max))

        return render_template('topsisaco.html', datosTopsisaco=datosTopsisaco)
    except Exception as e:
        return render_template('topsisaco.html', error_message=str(e))


@app.route('/topsisaco', methods=['POST'])
def calcular_topsisaco():
    try:
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos

        
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])
        
        # benefit_input = request.form['be']  # Obtén el valor
        
        # benefit_values = benefit_input.split(',')
        
        # benefit_attributes = [int(value) for value in benefit_values if value.strip() != '']

        # Llama a la función de PSO en pso.py
        datosTopsisaco = asyncio.run(ejecutar_topsisaco(w,alphaAco,beta,rho,Q,n_ants,iter_max))
        print("Resultados de la ejecución:", datosTopsisaco)

        # Devuelve los resultados como JSON
        return jsonify(datosTopsisaco)
    except Exception as e:
        # Manejo de errores
        print(f'Error en calcular_topsisaco: {str(e)}')
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------
@app.route('/comparacionGeneral')
def comparacionGeneral():
    try:
        # Generales
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos

        #Datos Pso
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Datos Ba
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        

        #Datos Aco
        ev_input = request.form['ev']  # Obtén el valor
        ev_values = ev_input.split(',')
        EV = [str(value) for value in ev_values if value.strip() != '']
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])
        
        #Llamar a funciones PSO
        datosDapso = asyncio.run(ejecutar_dapso(w,wwi,c1,c2,r1,r2 ,iter_max))
        datosMoorapso = asyncio.run(ejecutar_moorapso(w,wwi,c1,c2,r1,r2, iter_max))
        datosTopsispso = asyncio.run(ejecutar_topsispso(w,wwi,c1,c2,r1,r2 ,iter_max))
        # Llamar a funciones BA
        datosDaba = asyncio.run(ejecutar_daba(w, alpha, gamma, iter_max))
        datosMooraba = asyncio.run(ejecutar_mooraba(w, alpha, gamma, iter_max))
        datosTopsisba = asyncio.run(ejecutar_topsisba(w, alpha, gamma, iter_max))
        #Llamar a funciones ACO
        datosDaaco = asyncio.run(ejecutar_daaco(w, alphaAco, beta, rho, Q, n_ants, iter_max))
        datosMooraaco = asyncio.run(ejecutar_mooraaco(EV, w, alphaAco, beta, rho, Q, n_ants, iter_max))
        datosTopsisaco = asyncio.run(ejecutar_topsisaco(w, alphaAco, beta, iter_max))
        

        return render_template('comparacionGeneral.html', datosDapso=datosDapso, datosMoorapso=datosMoorapso, datosTopsispso=datosTopsispso,
                                datosDaba = datosDaba , datosMooraba = datosMooraba, datosTopsisba = datosTopsisba,
                                datosDaaco=datosDaaco, datosMooraaco=datosMooraaco, datosTopsisaco=datosTopsisaco)
    except Exception as e:
        return render_template('comparacionGeneral.html', error_message=str(e))


@app.route('/comparacionGeneral', methods=['POST'])
def calcular_comparacionGeneral():
    try:
        # Obtén los datos del formulario
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
       #Datos Pso
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Datos Ba
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
    

        #Datos Aco
        ev_input = request.form['ev']  # Obtén el valor
        ev_values = ev_input.split(',')
        EV = [str(value) for value in ev_values if value.strip() != '']
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])

         #Llamar a funciones PSO
        datosDapso = asyncio.run(ejecutar_dapso(w,wwi,c1,c2,r1,r2 ,iter_max))
        datosMoorapso = asyncio.run(ejecutar_moorapso(w,wwi,c1,c2,r1,r2 ,iter_max))
        datosTopsispso = asyncio.run(ejecutar_topsispso(w,wwi,c1,c2,r1,r2 ,iter_max))
        # Llamar a funciones BA
        datosDaba = asyncio.run(ejecutar_daba(w, alpha, gamma, iter_max))
        datosMooraba = asyncio.run(ejecutar_mooraba(w, alpha, gamma, iter_max))
        datosTopsisba = asyncio.run(ejecutar_topsisba(w, alpha, gamma, iter_max))
         #Llamar a funciones ACO
        datosDaaco = asyncio.run(ejecutar_daaco(w, alphaAco, beta, rho, Q, n_ants, iter_max))
        datosMooraaco = asyncio.run(ejecutar_mooraaco(EV, w, alphaAco, beta, rho, Q, n_ants, iter_max))
        datosTopsisaco = asyncio.run(ejecutar_topsisaco(w, alphaAco, beta, iter_max))
        
        
        print("Resultados de la ejecución:", datosDapso) 
        print("Resultados de la ejecución:", datosMoorapso) 
        print("Resultados de la ejecución:", datosTopsispso) 
        print("Resultados de la ejecución:", datosDaba) 
        print("Resultados de la ejecución:", datosMooraba) 
        print("Resultados de la ejecución:", datosTopsisba) 
        print("Resultados de la ejecución:", datosDaaco) 
        print("Resultados de la ejecución:", datosMooraaco) 
        print("Resultados de la ejecución:", datosTopsisaco) 

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosDapso,datosMoorapso,datosTopsispso, datosDaba, datosMooraba, datosTopsisba,datosDaaco,datosMooraaco,datosTopsisaco)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_comparacion: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500

#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------

@app.route('/comparacionBa')
def comparacionBa():
    try:
        # Obtén los datos del formulario
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos en pso.py
        datosBa = asyncio.run(ejecutar_ba(w, alpha, gamma, iter_max))
        datosDaba = asyncio.run(ejecutar_daba(w, alpha, gamma, iter_max))
        datosMooraba = asyncio.run(ejecutar_mooraba(w, alpha, gamma, iter_max))
        datosTopsisba = asyncio.run(ejecutar_topsisba(w, alpha, gamma, iter_max))
        
        
        

        return render_template('comparacionBa.html', datosBa=datosBa, datosDaba = datosDaba , datosMooraba = datosMooraba, datosTopsisba = datosTopsisba)
    except Exception as e:
        return render_template('comparacionBa.html', error_message=str(e))


@app.route('/comparacionBa', methods=['POST'])
def calcular_comparacionBa():
    try:
        # Obtén los datos del formulario
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alphaBa'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos en pso.py
        datosBa = asyncio.run(ejecutar_ba(w, alpha, gamma, iter_max))
        datosDaba = asyncio.run(ejecutar_daba(w, alpha, gamma, iter_max))
        datosMooraba = asyncio.run(ejecutar_mooraba(w, alpha, gamma, iter_max))
        datosTopsisba = asyncio.run(ejecutar_topsisba(w, alpha, gamma, iter_max))

        # Llama a la función de PSO en pso.py
        
        
        print("Resultados de la ejecución:", datosBa) 
        print("Resultados de la ejecución:", datosDaba) 
        print("Resultados de la ejecución:", datosMooraba) 
        print("Resultados de la ejecución:", datosTopsisba) 

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosBa, datosDaba, datosMooraba, datosTopsisba)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_comparacion: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500

#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------
# @app.route('/comparacionAco')
# def comparacionAco():
#     try:
#         # Obtén los datos del formulario
#         w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
#         w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
#         alpha = int(request.form['alpha'])
#         beta = int(request.form['beta'])
#         rho = float(request.form['rho'])
#         Q = int(request.form['Q'])
#         n_ants = int(request.form['n_ants'])
#         iter_max = int(request.form['iter_max'])
        
#         # Llama a la función de procesar_datos en pso.py
#         datosAco = asyncio.run(ejecutar_aco(w, alpha, beta, rho, Q, n_ants, iter_max))
#         datosDaaco = asyncio.run(ejecutar_daaco(w, alpha, beta, rho, Q, n_ants, iter_max))
#         datosMooraaco = asyncio.run(ejecutar_mooraaco(w, alpha, beta, iter_max))
#         datosTopsisaco = asyncio.run(ejecutar_topsisaco(w, alpha, beta, iter_max))
                

#         return render_template('comparacionAco.html', datosAco=datosAco, datosDaaco = datosDaaco , datosMooraaco = datosMooraaco, datosTopsisaco = datosTopsisaco)
#     except Exception as e:
#         return render_template('comparacionAco.html', error_message=str(e))


@app.route('/comparacionAco', methods=['GET', 'POST'])
def comparacionAco():
    if request.method == 'POST':
        try:
            ev_input = request.form['ev']  # Obtén el valor
            ev_values = ev_input.split(',')
            EV = [str(value) for value in ev_values if value.strip() != '']

             # Obtén los datos del formulario
            w_input = request.form['w']  # Obtén el valor seleccionado del menú desplegable desde el formulario

            # Divide la cadena en una lista de valores usando la coma como separador
            w_values = w_input.split(',')

            # Convierte cada valor en la lista a un número flotante, filtrando valores vacíos
            w = [float(value) for value in w_values if value.strip() != '']     
            alphaAco = int(request.form['alphaAco'])
            beta = int(request.form['beta'])
            rho = float(request.form['rho'])
            Q = int(request.form['Q'])
            n_ants = int(request.form['n_ants'])
            iter_max = int(request.form['T'])
            
            # Llama a la función de procesar_datos
            datosAco = asyncio.run(ejecutar_aco(w, alphaAco, beta, rho, Q, n_ants, iter_max))
            datosDaaco = asyncio.run(ejecutar_daaco(w, alphaAco, beta, rho, Q, n_ants, iter_max))
            datosMooraaco = asyncio.run(ejecutar_mooraaco(EV, w, alphaAco, beta, rho, Q, n_ants, iter_max))
            datosTopsisaco = asyncio.run(ejecutar_topsisaco(w, alphaAco, beta, iter_max))

            # Regresa los resultados como JSON
            return jsonify(datosAco=datosAco, datosDaaco=datosDaaco, datosMooraaco=datosMooraaco, datosTopsisaco=datosTopsisaco)
        except Exception as e:
            # Retorna un error JSON detallado
            return jsonify({'error': str(e)}), 500
    else:
        # Para el método GET, solo renderiza el template
        return render_template('comparacionAco.html')

#-------------------------------------------------------------------------------------------------------------------
        #Algoritmos MCDM - PUROS

#-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta ACO

@app.route('/aco')
def aco():
    try:
        # Obtén los datos del formulario
        w = [0.400, 0.200, 0.030, 0.070, 0.300]
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos en pso.py
        datosAco = asyncio.run(ejecutar_aco(w,alphaAco, beta, rho, Q, n_ants, iter_max))

        return render_template('aco.html', datosAco=datosAco)
    except Exception as e:
        return render_template('aco.html', error_message=str(e))


@app.route('/aco', methods=['POST'])
def calcular_aco():
    try:
       # Obtén los datos del formulario
        w = [0.400, 0.200, 0.030, 0.070, 0.300]
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos en pso.py
        datosAco = asyncio.run(ejecutar_aco(w, alphaAco, beta, rho, Q, n_ants, iter_max))
        print("Resultados de la ejecución:", datosAco)

        # Devuelve los resultados como JSON
        return jsonify(datosAco)
    except Exception as e:
        # Manejo de errores
        print(f'Error en calcular_aco: {str(e)}')
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------


#-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta TOPSIS

@app.route('/topsis')
def topsis():
    try:
        # Obtén los datos del formulario
        w_input = float(request.form.get('w', '')) 
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
       
        
        # Llama a la función de procesar_datos en pso.py
        datosTopsis = asyncio.run(ejecutar_topsis(w))

        return render_template('topsis.html', datosTopsis=datosTopsis)
    except Exception as e:
        return render_template('topsis.html', error_message=str(e))


@app.route('/topsis', methods=['POST'])
def calcular_topsis():
    try:
        # Obtén los datos del formulario
        w_input = request.form.get('w', '')  # Obtiene el valor de 'w' del formulario
        w_values = w_input.split(",")  # Divide la cadena en valores individuales
        w = [float(value.strip()) for value in w_values if value.strip()]  # Convierte cada valor a flotante

        # Llama a la función de PSO en pso.py
        datosTopsis = asyncio.run(ejecutar_topsis(w))
        print("Resultados de la ejecución:", datosTopsis)

        # Devuelve los resultados
        return jsonify(datosTopsis)
    except Exception as e:
        print(f'Error en calcular_topsis: {str(e)}')
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500


#-------------------------------------------------------------------------------------------------------------------


#-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta MOORAV

@app.route('/moorav')
def moorav():
    try:
        # Obtén los datos del formulario
        w_input = request.form.get('w', '')  # Obtiene el valor de 'w' del formulario
        w_values = w_input.split(",")  # Divide la cadena en valores individuales
        w = [float(value.strip()) for value in w_values if value.strip()]  # Convierte cada valor a flotante
        
        
        # Llama a la función de procesar_datos en pso.py
        datosMoorav = asyncio.run(ejecutar_moorav(w))

        return render_template('moorav.html', datosMoorav=datosMoorav)
    except Exception as e:
        return render_template('moorav.html', error_message=str(e))


@app.route('/moorav', methods=['POST'])
def calcular_moorav():
    try:
        # Obtén los datos del formulario
        w_input = request.form.get('w', '')  # Obtiene el valor de 'w' del formulario
        w_values = w_input.split(",")  # Divide la cadena en valores individuales
        w = [float(value.strip()) for value in w_values if value.strip()]  # Convierte cada valor a flotante
        

        # Llama a la función de PSO en pso.py
        datosMoorav = asyncio.run(ejecutar_moorav(w,))
        print("Resultados de la ejecución:", datosMoorav)

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosMoorav)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_moorav: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta MOORAV

@app.route('/da')
def da():
    try:
        # Obtén los datos del formulario
        w_input = request.form.get('w', '')  # Obtiene el valor de 'w' del formulario
        w_values = w_input.split(",")  # Divide la cadena en valores individuales
        w = [float(value.strip()) for value in w_values if value.strip()]  # Convierte cada valor a flotante
        
        
        # Llama a la función de procesar_datos en pso.py
        datosDa = asyncio.run(ejecutar_da(w))

        return render_template('da.html', datosDa=datosDa)
    except Exception as e:
        return render_template('da.html', error_message=str(e))


@app.route('/da', methods=['POST'])
def calcular_da():
    try:
        # Obtén los datos del formulario
        w_input = request.form.get('w', '')  # Obtiene el valor de 'w' del formulario
        w_values = w_input.split(",")  # Divide la cadena en valores individuales
        w = [float(value.strip()) for value in w_values if value.strip()]  # Convierte cada valor a flotante

        # Llama a la función de PSO en pso.py
        datosDa = asyncio.run(ejecutar_da(w))
        print("Resultados de la ejecución:", datosDa)

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosDa)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_Da: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------
@app.route('/index', methods=['POST'])
def index():
    # Parámetros de control (ingresan)
    w=request.form['w']
    #w=0.3
    # n=int(request.form['n'])  #partículas
    n=9
    c1=request.form['c1']
    #c1=0.50
    c2=request.form['c2']
    #c2=0.50
    e=int(request.form['e'])
    itera=e  #iteraciones

#Función objetivo
# Ri= (t^-i)/ (t^+i+t^-i) , i=1,...,m
# vector negativo / (vector positivo + vector negarivo)
# R2/(R2+R1)


    # Variables aleatorias, estas van a ser ingresadas(posteriormente)
    #R1 --> es el Vector positivo
    FILE_PATH='r1.xlsx'
    SHEET='Hoja1'
    workbook=load_workbook(FILE_PATH,read_only=True)
    sheet=workbook[SHEET]
    r1=[]
    for row in sheet.iter_rows():
        #print(row[0].value)
        r1.append(row[0].value)
        #print(r1)

    #R2 --> es el vector negativo
    #SHEET = 'Hoja1'
    FILE_PATH='r2.xlsx'
    workbook=load_workbook(FILE_PATH,read_only=True)
    sheet=workbook[SHEET]
    r2=[]
    for row in sheet.iter_rows():
        #print(row[0].value)
        r2.append(row[0].value)
        #print(r1)

    # ----------
    # Inicialización de la partícula del enjambre
    # Iniciando cálculo de las iteraciones
    print("")
    print("----")
    print("Iteración # 1")
    CP=[]
    for i in range(n):
        r11=float(r1[i])
        CP.append(float(format(10*(r11-0.5), '.4f')))
        #print("CP=",CP)
    #print("   CP=",CP)
    print("   Posición actual:       CP=",CP)

    # Inicialización de la velocidad
    V=[]
    for i in range(n):
        r21=float(r2[i])
        V.append(float(format((r21-0.5), '.4f')))
    #print("   V=",V)
    #print("   Velocidad actual:             V=",V)

    # Óptimo actual CF(1)
    # Ri= (t^-i)/ (t^+i+t^-i) , i=1,...,m
    # vector negativo / (vector positivo + vector negarivo)
    # R2/(R2+R1)
    CF=[]
    for i in range(n):
        #R2/(R1+R2)
        CForm1=float(r1[i])
        CForm2=float(r2[i])
        CForm3=(float((CForm2)/(CForm2+CForm1)))
        CForm=(format(float(CForm3), '.4f'))
        CF.append(float(CForm))
    print("   Óptimo actual:         CF=",CF)
    #print("   CF=",CF)

    # mejor posición actual
    LBP=[]
    for i in range(n):
        # LBP[1]=CP[1]
        LBPP=format(float(CP[i]), '.4f')
        LBP.append(float(LBPP))
    #print("   LBP=",LBP)
    print("   Mejor posición local:  LBP=",LBP)

    # mejor óptimo local
    LBF=[]
    for i in range(n):
        # LBF[1]=CF[1]
        LBFF=format(float(CF[i]), '.4f')
        LBF.append(float(LBFF))
    #print("   LBF=",LBF)
    print("   Mejor óptimo local:    LBF=",LBF)

    # MEJOR ÓPTIMO GLOBAL
    GBF=[]
    GBF.append(max(LBF))
    GBFt=max(LBF)
    #print("   GBF=", GBF)
    print("   Mejor óptimo global:   GBF=", GBF)

    # MEJOR POSICION GLOBAL
    GBP=[]
    GBF_index=LBF.index(GBFt)
    GBPt=float(LBP[GBF_index])
    GBP.append(LBP[GBF_index])
    #print("   GBP=",GBP)
    print("   Mejor posición global: GBP=",GBP)

    ###########################################
    #  Iteráción 2 a la n
    ##########################################
    it=2
    e=e-2

    while (e>=0):
        print("--------")
        print("Iteración #", it)
        long_V1=len(V)-n
        long_LBP1=len(LBP)-n
        long_CP1=len(CP)-n
        long_GBP1=len(GBP)-1
        for i in range(n):
            # V(i+1)= W*V(i) +c1*r1*(LBP(i)-CP(i))+c2*r2*(GBP)i)-CP(i))
            V1=format(float(w)*float(V[long_V1]), '.4f')
            V2=format(float(c1)*float(r1[i]), '.4f')
            V3=format(float(LBP[long_LBP1])-float(CP[long_CP1]), '.4f')
            V4=format(float(c2)*float(r2[i]), '.4f')
            V5=format(float(GBP[long_GBP1]), '.4f')
            V6=format(float(CP[long_CP1]), '.4f')
            V7=float(V5)-float(V6)
            Vx=format((float(V1)+float(V2)*float(V3)+float(V4)*float(V7)), '.4f')
            V.append(Vx)
            long_V1=long_V1+1
            long_LBP1=long_LBP1+1
            long_CP1=long_CP1+1
        long_V2=len(V)-n
        long_CP2=len(CP)-n

        #CP=CP(i)+V(i)
        for i in range(n):
            CPI=(format(float(CP[long_CP2])+float(V[long_V2]), '.4f'))
            CP.append(float(CPI))
            long_V2=long_V2+1
            long_CP2=long_CP2+1

        #óptimo actual CF
        long_CP1=len(CP)-(2*n)
        #print("long_CP2",long_CP2)
        long_CP2=len(CP)-n
        #print("long_CP3",long_CP3)
        for i in range(n):
            #CF2/(CF2+CF1)
            #print("CP",CP)
            CF1=CP[long_CP1]
            #print("CF1",CF1)
            CF2=CP[long_CP2]
            #print("CF2",CF2)
            CF3=float(CF2+CF1)
            CF12=format(float((CF2)/(CF3)),'.4f')
            CF.append(CF12)
            #print("CF12 ",CF12)
            long_CP1=long_CP1+1
            long_CP2=long_CP2+1

        # mejor óptimo local
        long_CF2=len(CF)-n
        long_LBF1=len(LBF)-(n)
        for i in range(n):
            #Max( CF[i],LBF[i-1])
            #print("CF",CF)
            #print("LBF",LBF)

            CFt=float(CF[long_CF2])
            #print("long_CF2",long_CF2)
            #print("CFt",CFt)
            
            LBFt=float(LBF[long_LBF1])
            #print("long_LBF1",long_LBF1)
            #print("LBFt",LBFt)
            if CFt>LBFt:
                LBF.append(CFt)
            else:
                LBF.append(LBFt)
            #print("LBF",LBF)
            long_CF2=long_CF2+1
            long_LBF1=long_LBF1+1

        # mejor posición actual
        long_CP4=len(CP)-n
        for i in range(n):
            #LBP[i]= posición de LBF[i]-CP[i]
            LBPt=(format(float(CP[long_CP4]), '.4f'))
            LBP.append(float(LBPt))
            long_CP4=long_CP4+1

        # MEJOR ÓPTIMO GLOBAL
        long_LBF=len(CP)-n
        temporal=[]
        for i in range(n):
            temporal.append(LBF[long_LBF])
            long_LBF=long_LBF+1
        GBF.append(max(temporal))
        GBFt=max(temporal)
        #print("GBFt=",GBFt)
        #print("MaxGlobal=",GBFt)

        # MEJOR POSICION GLOBAL
        #print("PosiciónGBP posición en CP")
        #print("CP=",CP)
        #long_GBF=len(CP)-n
        long_GBF=len(CP)-1
        #print("LBF",LBF)
        #print("long_CP=",long_GBF)
        GBPt=CP[long_GBF]
        #print("GBPt",GBPt)
        GBP.append(GBPt)


    # impresión de datos por iteración
        V_imp = len(V)-n
        CP_imp = len(CP)-n
        CF_imp = len(CF)-n
        LBF_imp = len(LBF)-n
        LBP_imp = len(LBP)-n
        GBF_imp = len(GBF)-1
        GBP_imp = len(GBP)-1

        V_impT = []
        for i in range(n):
            V_impT.append(float(V[V_imp]))
            V_imp = V_imp+1
        print("   V=",V_impT)

        CP_impT = []
        for i in range(n):
            CP_impT.append(float(CP[CP_imp]))
            CP_imp = CP_imp+1
        print("   CP=",CP_impT)

        CF_impT = []
        for i in range(n):
            CF_impT.append(float(CF[CF_imp]))
            CF_imp = CF_imp+1
        print("   CF=",CF_impT)

        LBF_impT = []
        for i in range(n):
            LBF_impT.append(float(LBF[LBF_imp]))
            LBF_imp = LBF_imp+1
        print("   LBF=",LBF_impT)

        LBP_impT = []
        for i in range(n):
            LBP_impT.append(float(LBP[LBP_imp]))
            LBP_imp = LBP_imp+1
        print("   LBP=",LBP_impT)

        for i in range(n):
            GBF_impT = (float(GBF[GBF_imp]))
        print("   GBF=",GBF_impT)

        for i in range(n):
            GBP_impT = (float(GBP[GBP_imp]))
        print("   GBP=",GBP_impT)

        e = e-1
        it = it+1
        print("")

    print("*******************")
    GBF_SF = len(GBF)-1
    GBP_SF = len(GBP)-1

    GBF_FIN = (float(GBF[GBF_SF]))
    print("   Mejor posición=",GBF_FIN)

    GBP_FIN = (float(GBP[GBP_SF]))
    print("   Mejor óptimo=",GBP_FIN)
    print("*******************")
    print("")

    context = {
        "c1": c1,
        "c2": c2,
        "w": w,
        "n": n,
        "e": e,
        "itera": itera,
        #"var1": var1,
        #"var2": var2,
        #"var3": var3,

        "r1": r1,
        "r2": r2,

        "V_impT": V_impT,
        "CP_impT": CP_impT,
        "CF_impT": CF_impT,
        "LBF_impT": LBF_impT,
        "LBP_impT": LBP_impT,
        "GBF_impT": GBF_impT,
        "GBP_impT": GBP_impT,

        "GBF_FIN": GBF_FIN,
        "GBP_FIN": GBP_FIN,
    }

    return render_template('index.html', **context)

@app.route('/')
def home():
    # if not session.get('logged_in'):
    #     return render_template('login.html')
    # else:
        return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def log_in():
    if request.method=='POST':
        emailLogin = request.form['email']
        passLogin = request.form['password']
        mensaje = ''
        password_Object = hashlib.sha256(passLogin.encode())
        passwordHashed =  password_Object.hexdigest()
        all_users = SupaUser.query.all()
        for user in all_users:
            print(f"ID: {user.id}, Username: {user.nombredb}, Email: {user.emaildb}, Password: {user.passworddb}")
        if emailLogin == user.emaildb and passwordHashed == user.passworddb:
            print("Usuario Correcto")
            mensaje = "Usuario Correcto"
            session['logged_in'] = True
            return render_template('index.html')
        else:
            print("Usuario Incorrecto")
            mensaje = "Usuario Incorrecto"
            return render_template('login.html')
    else:
        return render_template('login.html')

@app.route('/signup')
def sign_up():
    return render_template('signup.html')

@app.route('/signup', methods=['POST'])
def sign_validation():
    
    listaErrores = {
        'nombre': '',
        'escolaridad' : '',
        #'nacimiento' : '',
        'email' : '',
        'password' : ''
    }
    if request.method == 'POST': 
        nombre = request.form['nombre']
        escolaridad = request.form['escolaridad']
        #nacimiento = request.form['nacimiento']
        email = request.form['email']
        password = request.form['password']
        mensaje = ''
            
        if not nombre:
            listaErrores['nombre'] = 'Completa este campo'
        
        elif not re.match("^^[a-zA-Z\s]+$", nombre):
                listaErrores['nombre'] = 'Favor de ingresar solo letras'
        
        if not escolaridad:
            listaErrores['escolaridad'] = 'Completa este campo'
        
        elif not re.match("^[a-zA-Z\s]+$", escolaridad):
                listaErrores['escolaridad'] = 'Favor de ingresar solo letras'
        
        # if not nacimiento:
        #     listaErrores['nacimiento'] = 'Completa este campo'

        if not email:
            listaErrores['email'] = 'Completa este campo'
        
        elif not re.match(r'^[\w\.-]+@[\w\.-]+$', email):
                listaErrores['email'] = 'Ingresa un correo valido'
                
        if not password:
            listaErrores['password'] = 'Completa este campo'
                        
        elif len(password) < 8 or len(password) > 12 :
                listaErrores['password'] = 'La contrasena debe tener entre 8 y 12 caracteres' 
        
        if all(value == '' for value in listaErrores.values()):
            password_Object = hashlib.sha256(password.encode())
            passwordHashed =  password_Object.hexdigest()
            try:
                with app.app_context():
                    db.create_all()
                    user = SupaUser(nombredb=nombre, escolaridaddb=escolaridad, emaildb=email, passworddb=passwordHashed)
                    db.session.add(user)
                    db.session.commit()
                    mensaje = "Registro realizado con éxito!"
                return render_template("/login.html", mensaje=mensaje)
            except Exception as e:
                print(f'Error en la base de datos: {str(e)}')
                mensaje = "Ocurrió un error en la base de datos, por favor inténtalo de nuevo."
        else:
            return render_template("/signup.html", listaErrores=listaErrores)

    return render_template("/signup.html", listaErrores=listaErrores)


@app.route('/descargar-pso')
def descargar_excel_pso():
    directorio = 'Experimentos'  # Asegúrate de que este directorio exista y sea accesible
    filename = 'PSO.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)

@app.route('/descargar-dapso')
def descargar_excel_dapso():
    directorio = 'Experimentos'  # Asegúrate de que este directorio exista y sea accesible
    filename = 'DAPSO_1.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)

@app.route('/descargar-moorapso')
def descargar_excel_moorapso():
    directorio = 'Experimentos'  # Asegúrate de que este directorio exista y sea accesible
    filename = 'MOORAPSO_1.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)

@app.route('/descargar-topsispso')
def descargar_excel_topsispso():
    directorio = 'Experimentos'  # Asegúrate de que este directorio exista y sea accesible
    filename = 'TOPSISPSO_1.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)

@app.route('/descargar-ba')
def descargar_excel_ba():
    directorio = 'Experimentos'  # Asegúrate de que este directorio exista y sea accesible
    filename = 'BA.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)

@app.route('/descargar-daba')
def descargar_excel_daba():
    directorio = 'Experimentos'  # Asegúrate de que este directorio exista y sea accesible
    filename = 'DABA_1.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)

@app.route('/descargar-mooraba')
def descargar_excel_mooraba():
    directorio = 'Experimentos'  # Asegúrate de que este directorio exista y sea accesible
    filename = 'MOORABA_1.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)

@app.route('/descargar-topsisba')
def descargar_excel_topsisba():
    directorio = 'Experimentos'  # Asegúrate de que este directorio exista y sea accesible
    filename = 'TOPSISBA_1.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)

@app.route('/descargar-aco')
def descargar_excel_aco():
    directorio = 'Experimentos'  # Asegúrate de que este directorio exista y sea accesible
    filename = 'ACO_1.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)
#Aquí hubo un error
@app.route('/descargar-daaco')
def descargar_excel_daaco():
    directorio = 'Experimentos'  # Asegúrate de que este directorio exista y sea accesible
    filename = 'DAACO_1.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)

@app.route('/descargar-mooraaco')
def descargar_excel_mooraaco():
    directorio = 'Experimentos'  # Asegúrate de que este directorio exista y sea accesible
    filename = 'MOORAACO_1.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)

@app.route('/descargar-topsisaco')
def descargar_excel_topsisaco():
    directorio = 'Experimentos'  # Asegúrate de que este directorio exista y sea accesible
    filename = 'TOPSISACO_1.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)

@app.route('/descargar-topsis')
def descargar_excel_topsis():
    directorio = 'Experimentos'  # Asegúrate de que este directorio exista y sea accesible
    filename = 'TOPSIS_1.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)

@app.route('/descargar-moorav')
def descargar_excel_moorav():
    directorio = 'Experimentos'  # Asegúrate de que este directorio exista y sea accesible
    filename = 'MOORA_1.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)

@app.route('/descargar-da')
def descargar_excel_da():
    directorio = 'Experimentos'  # Asegúrate de que este directorio exista y sea accesible
    filename = 'DA_1.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)

@app.route('/descargar-zip')
def descargar_zip():
    directorio = ''  # Asegúrate de que este directorio exista y sea accesible
    filename = 'Compara.zip'
    return send_from_directory(directorio, filename, as_attachment=True)


@app.route('/descargar-parametros')
def descargar_parametros():
    directorio = ''  # Asegúrate de que este directorio exista y sea accesible
    filename = 'entradas-Programa.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)


if '__main__' == __name__:
    app.run(port=5000, debug=True)
